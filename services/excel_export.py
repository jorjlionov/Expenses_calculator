from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
import os


async def generate_excel(data: List[dict], user_id: int) -> str:
    """Генерирует Excel-файл со всеми транзакциями."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"

    # Заголовки
    headers = ["Тип", "Сумма", "Категория", "Дата"]
    ws.append(headers)

    # Стили заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Данные
    total_income = 0
    total_expense = 0

    for row in data:
        t_type = "Доход" if row["type"] == "income" else "Расход"
        amount = row["amount"]
        if row["type"] == "income":
            total_income += amount
        else:
            total_expense += amount

        ws.append([t_type, amount, row["category"], row["date"]])

    # Итоги
    ws.append([])
    ws.append(["ИТОГО", "", "", ""])
    ws.append(["Доходы:", total_income, "", ""])
    ws.append(["Расходы:", total_expense, "", ""])
    ws.append(["Баланс:", total_income - total_expense, "", ""])

    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    path = f"report_{user_id}.xlsx"
    wb.save(path)
    return path